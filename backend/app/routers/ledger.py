import io
import os
import re
import shutil
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel

from ..auth import get_current_user, require_roles, apply_mask, mask_value, get_client_ip
from ..config import ROLE_ADMIN, ROLE_MANAGER, ROLE_VIEWER, UPLOAD_DIR
from ..database import get_db, query_all, query_one, execute, loads, dumps, ensure_column
from ..services.audit import log_operation
from ..services.sync import notify_data_changed
from ..services import family

router = APIRouter(prefix="/api/ledger", tags=["ledger"])

SENSITIVE_FIELDS = {"id_card", "phone", "visa_no", "guardian_phone", "responsible_phone",
                    "parent_phone", "emergency_phone", "helper_phone"}

WRITABLE = [ROLE_ADMIN, ROLE_MANAGER]


def _calc_age_from_idcard(id_card):
    """按身份证号码计算年龄（支持18位/15位）"""
    if not id_card:
        return None
    s = str(id_card).strip()
    try:
        if len(s) == 18:
            b = s[6:14]
        elif len(s) == 15:
            b = "19" + s[6:12]
        else:
            return None
        birth = datetime.strptime(b, "%Y%m%d").date()
        today = date.today()
        age = today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
        return age if 0 <= age <= 150 else None
    except Exception:
        return None


def _calc_party_age(join_date):
    """按入党时间计算党龄（年，距今天数不足一年按 0）"""
    if not join_date:
        return None
    try:
        d = datetime.strptime(str(join_date)[:10], "%Y-%m-%d").date()
        today = date.today()
        years = today.year - d.year - ((today.month, today.day) < (d.month, d.day))
        return years if years >= 0 else None
    except Exception:
        return None


def _auto_fill(item: dict, field_map: dict) -> dict:
    """通用自动填充：台账含 age/party_age 字段且可推导时自动计算（增补优化）"""
    if "age" in field_map and item.get("id_card") and not item.get("age"):
        age = _calc_age_from_idcard(item.get("id_card"))
        if age is not None:
            item["age"] = age
    if "party_age" in field_map and item.get("join_date") and not item.get("party_age"):
        party_age = _calc_party_age(item.get("join_date"))
        if party_age is not None:
            item["party_age"] = party_age
    return item


def _menu(menu_code):
    m = query_one("SELECT * FROM sys_menu_config WHERE code=? AND is_ledger=1", (menu_code,))
    if not m:
        raise HTTPException(status_code=404, detail="台账不存在")
    return m


def _fields(menu_code, include_deleted=False):
    sql = "SELECT * FROM sys_field_config WHERE menu_code=?"
    params = [menu_code]
    if not include_deleted:
        sql += " AND is_deleted=0"
    rows = query_all(sql + " ORDER BY sort_order,id", params)
    out = []
    for f in rows:
        try:
            props = loads(f["props_json"] if f["props_json"] else None, {})
        except Exception:
            props = {}
        out.append({
            "id": f["id"], "physical_field": f["physical_field"], "display_label": f["display_label"],
            "data_type": f["data_type"], "form_component": f["form_component"], "is_system": f["is_system"],
            "show_in_list": f["show_in_list"], "show_in_form": f["show_in_form"],
            "is_required": f["is_required"], "sort_order": f["sort_order"],
            "options": loads(f["options_json"], []), "props": props,
        })
    return out


@router.get("/{menu_code}/duplicates")
def check_duplicates(menu_code: str, user: dict = Depends(get_current_user)):
    """户号/身份证号全局查重：返回重复数据，前端标红预警"""
    m = _menu(menu_code)
    fields = _fields(menu_code)
    checks = [f["physical_field"] for f in fields if f["physical_field"] in ("id_card", "household_no")]
    result = {}
    if "id_card" in checks:
        rows = query_all(
            f'SELECT "id_card" AS val, COUNT(*) c FROM {m["table_name"]} WHERE "id_card" IS NOT NULL AND "id_card"!=\'\' GROUP BY "id_card" HAVING c>1'
        )
        result["id_card"] = [{"value": r["val"], "count": r["c"]} for r in rows]
    if "household_no" in checks:
        rows = query_all(
            f'SELECT "household_no" AS val, COUNT(*) c FROM {m["table_name"]} WHERE "household_no" IS NOT NULL AND "household_no"!=\'\' GROUP BY "household_no" HAVING c>1'
        )
        result["household_no"] = [{"value": r["val"], "count": r["c"]} for r in rows]
    return result


@router.get("/{menu_code}/fields")
def ledger_fields(menu_code: str, user: dict = Depends(get_current_user)):
    m = _menu(menu_code)
    fields = _fields(menu_code)
    list_fields = [f for f in fields if f["show_in_list"]]
    form_fields = [f for f in fields if f["show_in_form"]]
    return {"menu": {"code": m["code"], "name": m["name"], "table": m["table_name"]},
            "list_fields": list_fields, "form_fields": form_fields, "fields": fields}


@router.get("/{menu_code}")
def list_data(menu_code: str, page: int = 1, size: int = 10, keyword: str = "",
              user: dict = Depends(get_current_user), request: Request = None):
    m = _menu(menu_code)
    fields = _fields(menu_code)
    filters = {}
    for key, val in request.query_params.items():
        if key.startswith("filter_"):
            filters[key[7:]] = val

    where = ["1=1"]
    params = []
    for f in fields:
        pf = f["physical_field"]
        if pf in filters and filters[pf] != "":
            where.append(f'"{pf}" = ?')
            params.append(filters[pf])
    if keyword:
        text_fields = [f["physical_field"] for f in fields if f["data_type"] in ("text", "select")]
        if text_fields:
            like = " OR ".join([f'"{pf}" LIKE ?' for pf in text_fields])
            where.append(f"({like})")
            for _ in text_fields:
                params.append(f"%{keyword}%")

    base = f'FROM {m["table_name"]} WHERE ' + " AND ".join(where)
    total = query_one(f"SELECT COUNT(*) c {base}", params)["c"]
    rows = query_all(
        f'SELECT * {base} ORDER BY id DESC LIMIT ? OFFSET ?',
        params + [size, (page - 1) * size],
    )
    list_fields = [f for f in fields if f["show_in_list"]]
    col_names = ["id"] + [f["physical_field"] for f in list_fields]
    data = [{k: r[k] for k in col_names if k in r} for r in rows]
    masked = apply_mask(data, [f["physical_field"] for f in list_fields])
    return {"total": total, "page": page, "size": size, "list": masked,
            "list_fields": list_fields}


@router.get("/{menu_code}/detail/{item_id}")
def item_detail(menu_code: str, item_id: int, user: dict = Depends(get_current_user)):
    m = _menu(menu_code)
    row = query_one(f'SELECT * FROM {m["table_name"]} WHERE id=?', (item_id,))
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    fields = _fields(menu_code)
    data = dict(row)
    if user["role"] == ROLE_VIEWER:
        for f in fields:
            if f["physical_field"] in SENSITIVE_FIELDS:
                data[f["physical_field"]] = mask_value(f["physical_field"], data.get(f["physical_field"]))
    return {"item": data, "fields": fields}


@router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(require_roles(*WRITABLE))):
    os.makedirs(IMAGE_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp",
                   ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".txt"):
        raise HTTPException(status_code=400, detail="仅支持图片或常用文档格式")
    name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{os.urandom(4).hex()}{ext}"
    path = os.path.join(IMAGE_DIR, name)
    with open(path, "wb") as out:
        shutil.copyfileobj(file.file, out)
    return {"url": f"/api/files/{name}"}


@router.get("/{menu_code}/household-check")
def household_check(menu_code: str, household_no: str = "", exclude_id: int = 0,
                    user: dict = Depends(get_current_user)):
    """查询户号下的家庭信息（成员、户主、户人数、是否单人户），供前端编辑/删除前校验与交接弹窗使用"""
    if not family.is_family_menu(menu_code):
        return {"enabled": False}
    with get_db() as db:
        members = family.family_members(db, household_no)
    out = [{"id": m["id"], "name": m.get("name"), "relation": m.get("relation"),
            "householder": m.get("householder")} for m in members
           if not exclude_id or m["id"] != exclude_id]
    holders = [m for m in out if m["householder"] == family.HOLDER]
    return {
        "enabled": True,
        "household_no": household_no or None,
        "size": len(members),
        "is_single": len(members) == 1,
        "holder": holders[0] if holders else None,
        "members": out,
    }


class TransferHouseholderBody(BaseModel):
    household_no: str
    current_holder_id: int
    new_holder_id: int


@router.post("/{menu_code}/transfer-householder")
async def transfer_householder(menu_code: str, body: TransferHouseholderBody,
                               user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    """户主交接：仅互换户主标志位，家庭人数与隶属关系不变"""
    if not family.is_family_menu(menu_code):
        raise HTTPException(status_code=400, detail="该台账不支持户主交接")
    m = _menu(menu_code)
    with get_db() as db:
        members = family.family_members(db, body.household_no)
        cur = [x for x in members if x["id"] == body.current_holder_id]
        nw = [x for x in members if x["id"] == body.new_holder_id]
        if not cur or not nw:
            raise HTTPException(status_code=400, detail="户主或新户主不在该户中")
        if cur[0].get("householder") != family.HOLDER:
            raise HTTPException(status_code=400, detail="当前成员不是户主，无需交接")
        if body.current_holder_id == body.new_holder_id:
            raise HTTPException(status_code=400, detail="户主不能交接给自己")
        db.execute(f'UPDATE {m["table_name"]} SET householder=? WHERE id=?',
                   (family.NOT_HOLDER, body.current_holder_id))
        db.execute(f'UPDATE {m["table_name"]} SET householder=? WHERE id=?',
                   (family.HOLDER, body.new_holder_id))
    log_operation(user, "户主交接", f"台账-{m['name']}",
                  f"户号{body.household_no} 户主交接 #{body.current_holder_id}->#{body.new_holder_id}",
                  get_client_ip(request))
    await notify_data_changed(menu_code)
    return {"message": "户主交接成功"}


@router.post("/{menu_code}")
async def create_item(menu_code: str, item: dict, user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    m = _menu(menu_code)
    fields = _fields(menu_code)
    field_map = {f["physical_field"]: f for f in fields}
    item = _auto_fill(item, field_map)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        if family.is_family_menu(menu_code) and item.get("household_no") not in (None, ""):
            if family.family_size(db, item["household_no"]) == 0 and "householder" not in item:
                item["householder"] = family.HOLDER
        cols, ph, vals = [], [], []
        for f in fields:
            pf = f["physical_field"]
            if pf in item and item[pf] is not None:
                cols.append(f'"{pf}"')
                ph.append("?")
                vals.append(item[pf])
        cols += ["create_time", "update_time"]
        ph += ["?", "?"]
        vals += [now, now]
        cur = db.execute(
            f'INSERT INTO {m["table_name"]} ({", ".join(cols)}) VALUES ({", ".join(ph)})', vals)
        new_id = cur.lastrowid
        if family.is_family_menu(menu_code):
            family.sync_population(db, item.get("household_no"))
    log_operation(user, "新增数据", f"台账-{m['name']}", f"新增记录#{new_id}", get_client_ip(request))
    await notify_data_changed(menu_code)
    return {"message": "新增成功", "id": new_id}


@router.put("/{menu_code}/{item_id}")
async def update_item(menu_code: str, item_id: int, item: dict, user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    m = _menu(menu_code)
    fields = _fields(menu_code)
    item = _auto_fill(item, {f["physical_field"]: f for f in fields})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        old_row = db.execute(f'SELECT * FROM {m["table_name"]} WHERE id=?', (item_id,)).fetchone()
        if not old_row:
            raise HTTPException(status_code=404, detail="记录不存在")
        old = dict(old_row)
        family_active = family.is_family_menu(menu_code)
        old_hno = old.get("household_no")
        new_hno = item.get("household_no")
        hno_changed = (str(new_hno) if new_hno is not None else "") != \
                      (str(old_hno) if old_hno is not None else "")
        old_holder = old.get("householder") == family.HOLDER
        new_holder = item.get("householder") == family.HOLDER
        if family_active and (hno_changed or (old_holder and not new_holder)):
            ok, msg, _ = family.guard_householder_change(db, old)
            if not ok:
                raise HTTPException(status_code=400, detail=msg)
        sets, vals = [], []
        for f in fields:
            pf = f["physical_field"]
            if pf in item:
                sets.append(f'"{pf}" = ?')
                vals.append(item[pf])
        sets.append("update_time = ?")
        vals.append(now)
        vals.append(item_id)
        cur = db.execute(f'UPDATE {m["table_name"]} SET {", ".join(sets)} WHERE id=?', vals)
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="记录不存在")
        if family_active:
            if new_holder:
                eff_hno = new_hno if new_hno is not None else old_hno
                for mb in family.family_members(db, eff_hno):
                    if mb["id"] != item_id and mb.get("householder") == family.HOLDER:
                        db.execute(f'UPDATE {m["table_name"]} SET householder=? WHERE id=?',
                                   (family.NOT_HOLDER, mb["id"]))
            if hno_changed:
                family.sync_population(db, old_hno)
            family.sync_population(db, new_hno)
    log_operation(user, "修改数据", f"台账-{m['name']}", f"修改记录#{item_id}", get_client_ip(request))
    await notify_data_changed(menu_code)
    return {"message": "修改成功"}


@router.delete("/{menu_code}/{item_id}")
async def delete_item(menu_code: str, item_id: int, user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    m = _menu(menu_code)
    with get_db() as db:
        old_row = db.execute(f'SELECT * FROM {m["table_name"]} WHERE id=?', (item_id,)).fetchone()
        if not old_row:
            raise HTTPException(status_code=404, detail="记录不存在")
        old = dict(old_row)
        if family.is_family_menu(menu_code):
            ok, msg, _ = family.guard_householder_change(db, old)
            if not ok:
                raise HTTPException(status_code=400, detail=msg)
        cur = db.execute(f"DELETE FROM {m['table_name']} WHERE id=?", (item_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="记录不存在")
        if family.is_family_menu(menu_code):
            family.sync_population(db, old.get("household_no"))
    log_operation(user, "删除数据", f"台账-{m['name']}", f"删除记录#{item_id}", get_client_ip(request))
    await notify_data_changed(menu_code)
    return {"message": "删除成功"}


IMAGE_DIR = os.path.join(UPLOAD_DIR, "images")


@router.get("/{menu_code}/export")
def export_excel(menu_code: str, tpl: str = "", user: dict = Depends(get_current_user), request: Request = None):
    from openpyxl import Workbook
    m = _menu(menu_code)
    fields = _fields(menu_code)
    if tpl:
        row = query_one("SELECT config_value FROM sys_config WHERE config_key=?", (f"export_tpl_{menu_code}",))
        tpl_fields = loads(row["config_value"] if row else None, None)
        if tpl_fields:
            fields = [f for f in fields if f["physical_field"] in tpl_fields]
        else:
            fields = [f for f in fields if f["show_in_list"]]
    else:
        fields = [f for f in fields if f["show_in_list"]]

    rows = query_all(f'SELECT * FROM {m["table_name"]} ORDER BY id')
    wb = Workbook()
    ws = wb.active
    ws.title = m["name"]
    headers = [f["display_label"] for f in fields]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(f["physical_field"], "") for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_operation(user, "导出数据", f"台账-{m['name']}", f"导出Excel {len(rows)}条", get_client_ip(request))
    from urllib.parse import quote
    fname = f"{m['name']}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


@router.post("/{menu_code}/import")
async def import_excel(menu_code: str, file: UploadFile = File(...),
                       user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    from openpyxl import load_workbook
    m = _menu(menu_code)
    fields = _fields(menu_code)
    label_map = {f["display_label"]: f["physical_field"] for f in fields}
    field_map = {f["physical_field"]: f for f in fields}
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件解析失败，请使用标准 .xlsx 文件")
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件中没有数据行")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = []
    for i, h in enumerate(headers):
        if h in label_map:
            col_map.append((i, label_map[h]))
    if not col_map:
        raise HTTPException(status_code=400, detail="表头无法匹配台账字段，请检查列名")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ok, dup = 0, 0
    with get_db() as db:
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            rec = {}
            for i, pf in col_map:
                rec[pf] = row[i]
            cols, ph, vals = [], [], []
            for pf, v in rec.items():
                cols.append(f'"{pf}"')
                ph.append("?")
                vals.append(v)
            cols += ["create_time", "update_time"]
            ph += ["?", "?"]
            vals += [now, now]
            # 身份证查重
            if "id_card" in rec and rec["id_card"]:
                exist = db.execute(f'SELECT id FROM {m["table_name"]} WHERE "id_card"=?', (rec["id_card"],)).fetchone()
                if exist:
                    dup += 1
                    continue
            try:
                db.execute(f'INSERT INTO {m["table_name"]} ({", ".join(cols)}) VALUES ({", ".join(ph)})', vals)
                ok += 1
            except Exception:
                continue
    log_operation(user, "导入数据", f"台账-{m['name']}", f"Excel导入 成功{ok}条 重复{dup}条", get_client_ip(request))
    await notify_data_changed(menu_code)
    return {"message": f"导入完成：成功 {ok} 条，身份证重复跳过 {dup} 条"}


class TplBody(BaseModel):
    fields: list


@router.get("/{menu_code}/templates")
def list_templates(menu_code: str, user: dict = Depends(get_current_user)):
    row = query_one("SELECT config_value FROM sys_config WHERE config_key=?", (f"export_tpl_{menu_code}",))
    if not row or not row["config_value"]:
        return {"templates": []}
    return {"templates": loads(row["config_value"], [])}


@router.post("/{menu_code}/templates")
async def save_template(menu_code: str, body: TplBody, user: dict = Depends(require_roles(*WRITABLE)), request: Request = None):
    fields = [f["physical_field"] for f in _fields(menu_code) if f["physical_field"] in body.fields]
    key = f"export_tpl_{menu_code}"
    with get_db() as db:
        db.execute("INSERT INTO sys_config(config_key,config_value,remark) VALUES(?,?,?) ON CONFLICT(config_key) DO UPDATE SET config_value=?",
                   (key, dumps(fields), "导出模板", dumps(fields)))
    log_operation(user, "保存导出模板", f"台账-{menu_code}", "保存字段导出模板", get_client_ip(request))
    return {"message": "模板已保存"}


@router.get("/{menu_code}/print/{item_id}")
def print_item(menu_code: str, item_id: int, user: dict = Depends(get_current_user)):
    m = _menu(menu_code)
    row = query_one(f'SELECT * FROM {m["table_name"]} WHERE id=?', (item_id,))
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    fields = [f for f in _fields(menu_code) if f["show_in_form"]]
    return {"menu_name": m["name"], "fields": fields, "item": dict(row)}
